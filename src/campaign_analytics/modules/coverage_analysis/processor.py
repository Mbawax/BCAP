"""Pure data processing engine for Coverage Analysis.

All public functions are free of Streamlit imports so they can be unit-tested
without a running Streamlit server.

Coverage Analysis combines three settlement-level datasets into a master table
and then aggregates to ward level.

Ward-Level Aggregation Method
-----------------------------
For each ward:
- % Visitation  = (count of visited settlements / total settlements in ward) × 100
- % Vaccination = mean of settlement-level Vaccination Coverage % values
- % Household Coverage = mean of settlement-level Household Coverage % values

The mean-of-settlement-percentages approach treats each settlement equally
regardless of population size.  This is the documented aggregation method.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import TypeAlias

import pandas as pd

MappingDict: TypeAlias = dict[str, str | None]

# Canonical field identifiers used by the column mapper.
VACCINATION_FIELDS = ("lga", "ward", "settlement", "vaccination_coverage")
HOUSEHOLD_FIELDS = ("lga", "ward", "settlement", "household_coverage")
VISITATION_FIELDS = ("lga", "ward", "settlement", "vaccination_coverage")
VISITATION_REQUIRED_FIELDS = ("lga", "ward", "settlement")

# Human-readable output column names.
COL_LGA = "LGA"
COL_WARD = "Ward"
COL_SETTLEMENT = "Settlement"
COL_VISITATION = "% Visitation"
COL_VACCINATION = "% Vaccination"
COL_HOUSEHOLD = "% Household Coverage"

SETTLEMENT_COLUMNS = [
    COL_LGA, COL_WARD, COL_SETTLEMENT,
    COL_VISITATION, COL_VACCINATION, COL_HOUSEHOLD,
]

WARD_COLUMNS = [
    COL_LGA, COL_WARD,
    "Total Settlements",
    "Visited Settlements",
    COL_VISITATION,
    COL_VACCINATION,
    COL_HOUSEHOLD,
]


# ── Data contracts ────────────────────────────────────────────────────────────

@dataclass(slots=True)
class ValidationReport:
    """Validation outcome for Coverage Analysis inputs."""

    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        """Return whether required data is valid for processing."""
        return not self.errors


@dataclass(frozen=True, slots=True)
class CoverageAnalysisSummary:
    """Campaign-level KPIs for Coverage Analysis."""

    total_settlements: int
    visited_settlements: int
    visitation_pct: float
    avg_vaccination_pct: float
    avg_household_pct: float
    total_wards: int
    total_lgas: int


@dataclass(frozen=True, slots=True)
class CoverageThreshold:
    """Configurable Red/Yellow/Green threshold for a single coverage indicator.

    Classification logic:
    - Red:    value < yellow_min
    - Yellow: yellow_min <= value < green_min
    - Green:  value >= green_min

    ``yellow_min`` is the boundary between Red and Yellow.
    ``green_min`` is the boundary between Yellow and Green.
    """

    yellow_min: float
    green_min: float

    def classify(self, value: float | None) -> str:
        """Return '🔴 Red', '🟡 Yellow', '🟢 Green', or '⚪ N/A' for a percentage value."""
        if value is None or pd.isna(value):
            return "⚪ N/A"
        if value < self.yellow_min:
            return "🔴 Red"
        if value < self.green_min:
            return "🟡 Yellow"
        return "🟢 Green"


# Default thresholds — used when the user has not configured custom values.
DEFAULT_THRESHOLDS: dict[str, CoverageThreshold] = {
    COL_VISITATION: CoverageThreshold(yellow_min=49.0, green_min=70.0),
    COL_VACCINATION: CoverageThreshold(yellow_min=49.0, green_min=70.0),
    COL_HOUSEHOLD: CoverageThreshold(yellow_min=49.0, green_min=70.0),
}

# Status column suffixes appended for each indicator.
STATUS_VISITATION = "Visitation Status"
STATUS_VACCINATION = "Vaccination Status"
STATUS_HOUSEHOLD = "Household Status"

_INDICATOR_STATUS_MAP: dict[str, str] = {
    COL_VISITATION: STATUS_VISITATION,
    COL_VACCINATION: STATUS_VACCINATION,
    COL_HOUSEHOLD: STATUS_HOUSEHOLD,
}


# ── Column suggestion helpers ─────────────────────────────────────────────────

def _normalise_name(name: str) -> str:
    """Normalise a column header to lowercase alphanumeric for matching."""
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def suggest_vaccination_mapping(columns: list[str]) -> MappingDict:
    """Suggest column mappings for Vaccination Coverage dataset."""
    return _suggest_mapping(
        columns,
        {
            "lga": ("lga", "localgovernmentarea", "district", "lganame"),
            "ward": ("ward", "wardname", "ward_name", "subdistrict"),
            "settlement": (
                "settlement", "settlementname", "settlement_name",
                "community", "village",
            ),
            "vaccination_coverage": (
                "vaccinationcoverage", "coveragepct", "coverage",
                "vaccinationcoveragepct", "vacccoverage",
                "vaccination_coverage", "coveragepercent",
            ),
        },
    )


def suggest_household_mapping(columns: list[str]) -> MappingDict:
    """Suggest column mappings for Household Coverage dataset."""
    return _suggest_mapping(
        columns,
        {
            "lga": ("lga", "localgovernmentarea", "district", "lganame"),
            "ward": ("ward", "wardname", "ward_name", "subdistrict"),
            "settlement": (
                "settlement", "settlementname", "settlement_name",
                "community", "village",
            ),
            "household_coverage": (
                "householdcoverage", "householdcoveragepct",
                "hhcoverage", "hhcoveragepct", "household_coverage",
                "householdcoveragepercent",
            ),
        },
    )


def suggest_visitation_mapping(columns: list[str]) -> MappingDict:
    """Suggest column mappings for Settlement Visitation dataset."""
    return _suggest_mapping(
        columns,
        {
            "lga": ("lga", "localgovernmentarea", "district", "lganame"),
            "ward": ("ward", "wardname", "ward_name", "subdistrict"),
            "settlement": (
                "settlement", "settlementname", "settlement_name",
                "community", "village",
            ),
            "vaccination_coverage": (
                "vaccinationcoverage", "coveragepct", "coverage",
                "vaccinationcoveragepct", "vacccoverage",
                "vaccination_coverage", "coveragepercent",
            ),
        },
    )


def _suggest_mapping(
    columns: list[str], aliases: dict[str, tuple[str, ...]]
) -> MappingDict:
    """Auto-suggest column mappings from a dictionary of normalised aliases."""
    normalised = {_normalise_name(col): col for col in columns}
    return {
        field_key: next(
            (normalised[name] for name in names if name in normalised), None
        )
        for field_key, names in aliases.items()
    }


# ── Text normalisation ────────────────────────────────────────────────────────

def normalize_text_series(series: pd.Series) -> pd.Series:
    """Clean and standardize text fields for exact key creation."""
    return series.fillna("").astype(str).str.strip().str.upper()


def build_unique_id(
    df: pd.DataFrame, lga_col: str, ward_col: str, settlement_col: str
) -> pd.Series:
    """Create a composite unique identifier (LGA_Ward_Settlement) for joining."""
    lga = normalize_text_series(df[lga_col])
    ward = normalize_text_series(df[ward_col])
    settlement = normalize_text_series(df[settlement_col])
    return lga + "_" + ward + "_" + settlement


# ── Validation ────────────────────────────────────────────────────────────────

def validate_inputs(
    df_vaccination: pd.DataFrame | None,
    vaccination_mapping: MappingDict,
    df_household: pd.DataFrame | None = None,
    household_mapping: MappingDict | None = None,
    df_visitation: pd.DataFrame | None = None,
    visitation_mapping: MappingDict | None = None,
) -> ValidationReport:
    """Validate uploaded dataframes and field mappings for Coverage Analysis."""
    report = ValidationReport()

    # Vaccination Coverage is the baseline — required.
    if df_vaccination is None or df_vaccination.empty:
        report.errors.append(
            "Vaccination Coverage dataset is required as the baseline settlement list."
        )
        return report

    # Validate vaccination mapping (all four fields required).
    _validate_required_mapping(
        df_vaccination, vaccination_mapping, VACCINATION_FIELDS,
        "Vaccination Coverage", report,
    )

    # Validate household mapping if uploaded.
    if df_household is not None and not df_household.empty and household_mapping:
        _validate_required_mapping(
            df_household, household_mapping, HOUSEHOLD_FIELDS,
            "Household Coverage", report,
        )
        # Validate numeric column.
        hh_col = household_mapping.get("household_coverage")
        if hh_col and hh_col in df_household.columns:
            _validate_percentage_column(
                df_household, hh_col, "Household Coverage %", report,
            )

    # Validate visitation mapping if uploaded.
    if df_visitation is not None and not df_visitation.empty and visitation_mapping:
        _validate_required_mapping(
            df_visitation, visitation_mapping, VISITATION_REQUIRED_FIELDS,
            "Settlement Visitation", report,
        )
        vis_vacc_col = visitation_mapping.get("vaccination_coverage")
        if vis_vacc_col and vis_vacc_col in df_visitation.columns:
            _validate_percentage_column(
                df_visitation, vis_vacc_col, "Settlement Visitation Vaccination Coverage %", report,
            )

    # Validate vaccination coverage numeric column.
    vacc_col = vaccination_mapping.get("vaccination_coverage")
    if vacc_col and vacc_col in df_vaccination.columns:
        _validate_percentage_column(
            df_vaccination, vacc_col, "Vaccination Coverage %", report,
        )

    # Advisory warnings when optional datasets are missing.
    if df_household is None or df_household.empty:
        report.warnings.append(
            "No Household Coverage dataset uploaded. "
            "Household Coverage % will show as 'N/A' in the master table."
        )
    if df_visitation is None or df_visitation.empty:
        report.warnings.append(
            "No Settlement Visitation dataset uploaded. "
            "% Visitation will show as 0% for all settlements."
        )

    return report


def _validate_required_mapping(
    df: pd.DataFrame,
    mapping: MappingDict,
    fields: tuple[str, ...],
    dataset_name: str,
    report: ValidationReport,
) -> None:
    """Check that every required field is mapped and the column exists."""
    for field_key in fields:
        mapped_col = mapping.get(field_key)
        if not mapped_col:
            report.errors.append(
                f"{dataset_name}: please map the required field "
                f"'{field_key.replace('_', ' ').title()}'."
            )
        elif mapped_col not in df.columns:
            report.errors.append(
                f"{dataset_name}: mapped column '{mapped_col}' "
                f"for '{field_key.replace('_', ' ').title()}' was not found in the file."
            )


def _validate_percentage_column(
    df: pd.DataFrame,
    column: str,
    label: str,
    report: ValidationReport,
) -> None:
    """Check a percentage column for blanks and out-of-range values."""
    numeric = _coerce_numeric(df[column])
    invalid_count = int(numeric.isna().sum())
    if invalid_count:
        report.warnings.append(
            f"{label}: {invalid_count:,} blank or non-numeric value(s) "
            "will be treated as N/A."
        )
    negative = int((numeric.dropna() < 0).sum())
    if negative:
        report.warnings.append(
            f"{label}: {negative:,} value(s) below 0% found."
        )
    over_100 = int((numeric.dropna() > 100).sum())
    if over_100:
        report.warnings.append(
            f"{label}: {over_100:,} value(s) above 100% found."
        )


def _coerce_numeric(values: pd.Series) -> pd.Series:
    """Coerce a column to numeric percentage values.

    Handles:
    - Strings with '%' like '60%' -> 60.0
    - Strings with commas like '1,000' -> 1000.0
    - Excel decimal percentages like 0.60 or 0.85 -> 60.0 or 85.0
    - Standard numbers like 60 or 60.0 -> 60.0
    """
    if values is None or values.empty:
        return pd.Series(dtype=float)

    s_str = (
        values.astype("string")
        .str.replace("%", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.strip()
    )
    numeric = pd.to_numeric(s_str, errors="coerce")

    valid_nums = numeric.dropna()
    if not valid_nums.empty:
        non_zero = valid_nums[valid_nums != 0]
        if not non_zero.empty and (non_zero.abs() <= 1.0).all() and (non_zero.max() > 0):
            numeric = numeric * 100.0

    return numeric.round(1)


# ── Processing ────────────────────────────────────────────────────────────────

def build_settlement_master(
    df_vaccination: pd.DataFrame,
    vaccination_mapping: MappingDict,
    df_household: pd.DataFrame | None = None,
    household_mapping: MappingDict | None = None,
    df_visitation: pd.DataFrame | None = None,
    visitation_mapping: MappingDict | None = None,
) -> tuple[CoverageAnalysisSummary, pd.DataFrame]:
    """Build the settlement-level master table from three input datasets.

    The Vaccination Coverage dataset provides the baseline settlement roster.
    Household Coverage and Settlement Visitation are left-joined onto it via
    a composite key (LGA_WARD_SETTLEMENT, uppercased and stripped).

    Returns:
        (summary_kpis, settlement_master_dataframe)
    """
    v_lga = vaccination_mapping["lga"]
    v_ward = vaccination_mapping["ward"]
    v_settlement = vaccination_mapping["settlement"]
    v_coverage = vaccination_mapping["vaccination_coverage"]

    # Build baseline from vaccination dataset.
    master = pd.DataFrame({
        COL_LGA: df_vaccination[v_lga].astype(str).str.strip(),
        COL_WARD: df_vaccination[v_ward].astype(str).str.strip(),
        COL_SETTLEMENT: df_vaccination[v_settlement].astype(str).str.strip(),
        COL_VACCINATION: _coerce_numeric(df_vaccination[v_coverage]),
    })
    master["unique_id"] = build_unique_id(
        df_vaccination, v_lga, v_ward, v_settlement
    )

    # ── Join Household Coverage ───────────────────────────────────────────
    if (
        df_household is not None
        and not df_household.empty
        and household_mapping
        and all(household_mapping.get(f) for f in HOUSEHOLD_FIELDS)
    ):
        h_lga = household_mapping["lga"]
        h_ward = household_mapping["ward"]
        h_settlement = household_mapping["settlement"]
        h_coverage = household_mapping["household_coverage"]

        hh = pd.DataFrame({
            "unique_id": build_unique_id(
                df_household, h_lga, h_ward, h_settlement
            ),
            COL_HOUSEHOLD: _coerce_numeric(df_household[h_coverage]),
        })
        # Deduplicate on key — take the first occurrence.
        hh = hh.drop_duplicates(subset=["unique_id"], keep="first")
        master = master.merge(hh, on="unique_id", how="left")
    else:
        master[COL_HOUSEHOLD] = pd.NA

    # ── Join Settlement Visitation ────────────────────────────────────────
    if (
        df_visitation is not None
        and not df_visitation.empty
        and visitation_mapping
        and all(visitation_mapping.get(f) for f in VISITATION_REQUIRED_FIELDS)
    ):
        vis_lga = visitation_mapping["lga"]
        vis_ward = visitation_mapping["ward"]
        vis_settlement = visitation_mapping["settlement"]

        vis_ids = build_unique_id(df_visitation, vis_lga, vis_ward, vis_settlement)

        # Check if a coverage column is mapped in visitation dataset
        vis_cov_col = (
            visitation_mapping.get("visitation_coverage")
            or visitation_mapping.get("vaccination_coverage")
        )

        if vis_cov_col and vis_cov_col in df_visitation.columns:
            vis_df = pd.DataFrame({
                "unique_id": vis_ids,
                "_vis_pct": _coerce_numeric(df_visitation[vis_cov_col]),
            }).drop_duplicates(subset=["unique_id"], keep="first")

            master = master.merge(vis_df, on="unique_id", how="left")
            in_visitation = master["unique_id"].isin(set(vis_ids))
            master[COL_VISITATION] = master["_vis_pct"].fillna(
                in_visitation.map({True: 100.0, False: 0.0})
            )
            master = master.drop(columns=["_vis_pct"])
        else:
            visited_set = set(vis_ids)
            master[COL_VISITATION] = master["unique_id"].isin(visited_set).map(
                {True: 100.0, False: 0.0}
            )
    else:
        master[COL_VISITATION] = 0.0

    # Drop internal join key.
    master = master.drop(columns=["unique_id"])

    # Round percentage columns to one decimal.
    for col in [COL_VACCINATION, COL_HOUSEHOLD]:
        if col in master.columns:
            master[col] = pd.to_numeric(master[col], errors="coerce").round(1)

    # Reorder columns to match the spec.
    master = master[SETTLEMENT_COLUMNS].reset_index(drop=True)

    # ── Summary KPIs ─────────────────────────────────────────────────────
    total_settlements = len(master)
    visited = int((master[COL_VISITATION] == 100.0).sum())
    visitation_pct = round(visited / total_settlements * 100, 1) if total_settlements > 0 else 0.0
    avg_vacc = round(master[COL_VACCINATION].dropna().mean(), 1) if master[COL_VACCINATION].notna().any() else 0.0
    avg_hh = round(master[COL_HOUSEHOLD].dropna().mean(), 1) if master[COL_HOUSEHOLD].notna().any() else 0.0
    total_wards = master[COL_WARD].nunique()
    total_lgas = master[COL_LGA].nunique()

    summary = CoverageAnalysisSummary(
        total_settlements=total_settlements,
        visited_settlements=visited,
        visitation_pct=visitation_pct,
        avg_vaccination_pct=avg_vacc,
        avg_household_pct=avg_hh,
        total_wards=total_wards,
        total_lgas=total_lgas,
    )

    return summary, master


def build_ward_summary(df_settlement: pd.DataFrame) -> pd.DataFrame:
    """Aggregate the settlement master table to ward level.

    Aggregation method (documented):
    - Total Settlements: count of settlements in the ward
    - Visited Settlements: count where % Visitation == 100
    - % Visitation: (Visited / Total) × 100
    - % Vaccination: mean of settlement-level Vaccination Coverage % values
    - % Household Coverage: mean of settlement-level Household Coverage % values

    The mean-of-settlement-percentages approach treats each settlement equally
    regardless of population size.
    """
    ward_groups = df_settlement.groupby([COL_LGA, COL_WARD], dropna=False)

    ward_rows: list[dict] = []
    for (lga, ward), group in ward_groups:
        total = len(group)
        visited = int((group[COL_VISITATION] == 100.0).sum())
        vis_pct = round(visited / total * 100, 1) if total > 0 else 0.0

        vacc_mean = group[COL_VACCINATION].dropna().mean()
        vacc_pct = round(vacc_mean, 1) if pd.notna(vacc_mean) else None

        hh_mean = group[COL_HOUSEHOLD].dropna().mean()
        hh_pct = round(hh_mean, 1) if pd.notna(hh_mean) else None

        ward_rows.append({
            COL_LGA: lga,
            COL_WARD: ward,
            "Total Settlements": total,
            "Visited Settlements": visited,
            COL_VISITATION: vis_pct,
            COL_VACCINATION: vacc_pct,
            COL_HOUSEHOLD: hh_pct,
        })

    df_ward = pd.DataFrame(ward_rows, columns=WARD_COLUMNS)
    return df_ward.sort_values([COL_LGA, COL_WARD]).reset_index(drop=True)


def apply_thresholds(
    df: pd.DataFrame,
    thresholds: dict[str, CoverageThreshold],
) -> pd.DataFrame:
    """Add Red/Yellow/Green status columns to a DataFrame based on thresholds.

    This is a **presentation-only** layer.  The underlying percentage columns
    are never modified.  A new status column is appended for each indicator
    that exists in the DataFrame.

    Args:
        df: Settlement-level or ward-level DataFrame containing indicator columns.
        thresholds: Mapping of indicator column name to its CoverageThreshold.

    Returns:
        A copy of the DataFrame with status columns appended.
    """
    result = df.copy()
    for indicator_col, status_col in _INDICATOR_STATUS_MAP.items():
        if indicator_col in result.columns and indicator_col in thresholds:
            threshold = thresholds[indicator_col]
            result[status_col] = result[indicator_col].apply(threshold.classify)
    return result


# ── Operational Issue Detection ───────────────────────────────────────────────

COL_ISSUE_TYPE = "Issue Type"

# Predefined issue type catalogue.  Each entry is a tuple of
# (issue_label, condition_function) where the condition function receives a row
# (with status columns) and returns True if the issue applies.
#
# "High" = Green, "Medium" = Yellow, "Low" = Red for threshold-derived status.
# Conditions test status columns, not raw percentages, so they automatically
# respect whatever thresholds the user has configured.

ISSUE_TYPES: list[tuple[str, str]] = [
    (
        "High Visitation + Low Vaccination",
        "Settlements were visited but vaccination coverage remains low.",
    ),
    (
        "High Visitation + Low Household Coverage",
        "Settlements were visited but household coverage remains low.",
    ),
    (
        "High Vaccination + Low Household Coverage",
        "Vaccination coverage is high but household coverage is low.",
    ),
    (
        "Low Visitation + High Vaccination",
        "Vaccination is high despite low visitation — possible data quality issue.",
    ),
    (
        "Low Visitation + High Household Coverage",
        "Household coverage is high despite low visitation — possible data quality issue.",
    ),
    (
        "Low Vaccination + High Household Coverage",
        "Household coverage is high but vaccination coverage is low.",
    ),
    (
        "Any Indicator Below Threshold",
        "At least one indicator is classified as Red.",
    ),
    (
        "Multiple Issues",
        "Ward has more than one misalignment pattern.",
    ),
]

ISSUE_LABELS = [label for label, _ in ISSUE_TYPES]
ISSUE_DESCRIPTIONS = {label: desc for label, desc in ISSUE_TYPES}


def _classify_ward_issues(row: pd.Series) -> list[str]:
    """Classify a single ward row into zero or more issue types.

    Uses the status columns (Visitation Status, Vaccination Status,
    Household Status) which are derived from the user's configured thresholds.
    """
    vis = str(row.get(STATUS_VISITATION, ""))
    vacc = str(row.get(STATUS_VACCINATION, ""))
    hh = str(row.get(STATUS_HOUSEHOLD, ""))

    is_green = lambda s: "Green" in s
    is_red = lambda s: "Red" in s

    issues: list[str] = []

    # 1. High Visitation + Low Vaccination
    if is_green(vis) and is_red(vacc):
        issues.append("High Visitation + Low Vaccination")

    # 2. High Visitation + Low Household Coverage
    if is_green(vis) and is_red(hh):
        issues.append("High Visitation + Low Household Coverage")

    # 3. High Vaccination + Low Household Coverage
    if is_green(vacc) and is_red(hh):
        issues.append("High Vaccination + Low Household Coverage")

    # 4. Low Visitation + High Vaccination
    if is_red(vis) and is_green(vacc):
        issues.append("Low Visitation + High Vaccination")

    # 5. Low Visitation + High Household Coverage
    if is_red(vis) and is_green(hh):
        issues.append("Low Visitation + High Household Coverage")

    # 6. Low Vaccination + High Household Coverage
    if is_red(vacc) and is_green(hh):
        issues.append("Low Vaccination + High Household Coverage")

    # 7. Any Indicator Below Threshold (Red)
    if any(is_red(s) for s in [vis, vacc, hh]):
        if "Any Indicator Below Threshold" not in issues:
            issues.append("Any Indicator Below Threshold")

    return issues


def detect_operational_issues(
    df_ward: pd.DataFrame,
    thresholds: dict[str, CoverageThreshold],
) -> pd.DataFrame:
    """Detect operational misalignment issues in ward-level data.

    Applies thresholds to produce status columns, then classifies each ward
    into zero or more issue types.  Wards with multiple specific issues
    (excluding the catch-all 'Any Indicator Below Threshold') are additionally
    tagged as 'Multiple Issues'.

    Returns a DataFrame with the standard ward columns, status columns, and an
    'Issue Type' column.  Wards with multiple issues appear once per issue type.
    Wards with no detected issues are excluded.
    """
    # Ensure status columns exist.
    df_with_status = apply_thresholds(df_ward, thresholds)

    issue_rows: list[dict] = []
    for _, row in df_with_status.iterrows():
        issues = _classify_ward_issues(row)

        # Count specific issues (excluding the catch-all).
        specific = [i for i in issues if i != "Any Indicator Below Threshold"]
        if len(specific) > 1:
            issues.append("Multiple Issues")

        for issue in issues:
            issue_row = row.to_dict()
            issue_row[COL_ISSUE_TYPE] = issue
            issue_rows.append(issue_row)

    if not issue_rows:
        # Return an empty DataFrame with the expected schema.
        cols = list(df_with_status.columns) + [COL_ISSUE_TYPE]
        return pd.DataFrame(columns=cols)

    df_issues = pd.DataFrame(issue_rows)

    # Reorder columns: ward identifiers, settlements, indicators, statuses, issue.
    desired_order = [
        COL_LGA, COL_WARD, "Total Settlements",
        COL_VISITATION, COL_VACCINATION, COL_HOUSEHOLD,
        STATUS_VISITATION, STATUS_VACCINATION, STATUS_HOUSEHOLD,
        COL_ISSUE_TYPE,
    ]
    available = [c for c in desired_order if c in df_issues.columns]
    extra = [c for c in df_issues.columns if c not in desired_order]
    df_issues = df_issues[available + extra].reset_index(drop=True)

    return df_issues


# ── Excel export ──────────────────────────────────────────────────────────────

def _apply_excel_formatting(
    ws: object,
    sheet_name: str,
    thresholds: dict[str, CoverageThreshold] | None = None,
) -> None:
    """Apply openpyxl styling, cell background colors, percentage formatting, and centering to Excel sheets.

    - Percentage indicator columns (% Visitation, % Vaccination, % Household Coverage):
      - Centered text alignment
      - Formatted cleanly with '%' suffix (e.g. 60%, 100%, 60.5%)
      - Background fills matching pipeline UI:
        - >=100%: Green (#BBF7D0 fill, #166534 text)
        - <100%: Classified via active RAG thresholds (Red/Yellow/Green/N/A)
    - Status & count columns:
      - Centered text alignment
    - Header row:
      - Navy Blue background (#1E3A8A), White bold font, centered alignment
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    font_header = Font(name="Calibri", size=11, color="FFFFFF", bold=True)

    fill_green = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")
    font_green = Font(name="Calibri", size=11, color="166534", bold=True)

    fill_yellow = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    font_yellow = Font(name="Calibri", size=11, color="854D0E", bold=True)

    fill_red = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    font_red = Font(name="Calibri", size=11, color="991B1B", bold=True)

    fill_na = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    font_na = Font(name="Calibri", size=11, color="6B7280")

    fill_issue = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    font_issue = Font(name="Calibri", size=11, color="854D0E", bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    border_thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    indicator_cols = {COL_VISITATION, COL_VACCINATION, COL_HOUSEHOLD}
    status_cols = {STATUS_VISITATION, STATUS_VACCINATION, STATUS_HOUSEHOLD}
    count_cols = {"Total Settlements", "Visited Settlements"}

    # Format header row
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    active_thresholds = thresholds or DEFAULT_THRESHOLDS

    # Format data rows
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            col_name = str(headers[col_idx - 1])

            if col_name in count_cols or col_name in status_cols:
                cell.alignment = align_center

            elif col_name in indicator_cols:
                cell.alignment = align_center
                val = cell.value
                if val is None or pd.isna(val) or str(val).strip() == "" or str(val).upper() == "N/A":
                    cell.value = "N/A"
                    cell.fill = fill_na
                    cell.font = font_na
                else:
                    try:
                        num = float(val)
                        if num == int(num):
                            cell.value = f"{int(num)}%"
                        else:
                            cell.value = f"{num:.1f}%"

                        if num >= 100.0:
                            cell.fill = fill_green
                            cell.font = font_green
                        else:
                            t = active_thresholds.get(col_name)
                            if t:
                                st_str = t.classify(num)
                                if "Red" in st_str:
                                    cell.fill = fill_red
                                    cell.font = font_red
                                elif "Yellow" in st_str:
                                    cell.fill = fill_yellow
                                    cell.font = font_yellow
                                else:
                                    cell.fill = fill_green
                                    cell.font = font_green
                            else:
                                cell.fill = fill_green
                                cell.font = font_green
                    except (ValueError, TypeError):
                        pass

            elif col_name == COL_ISSUE_TYPE:
                cell.alignment = align_left
                cell.fill = fill_issue
                cell.font = font_issue

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)


def generate_coverage_excel_report(
    summary: CoverageAnalysisSummary | None,
    df_settlement: pd.DataFrame,
    df_ward: pd.DataFrame,
    thresholds: dict[str, CoverageThreshold] | None = None,
    df_issues: pd.DataFrame | None = None,
) -> bytes:
    """Generate multi-sheet Executive Summary Excel workbook for Coverage Analysis.

    Sheets:
    1. Executive Summary — high-level KPIs and campaign overview
    2. Settlement Coverage — full settlement master table with RAG styling
    3. Ward Coverage — ward-aggregated indicators with RAG styling
    4. Operational Issues — detected misalignment issues (if any) with RAG styling
    5. Threshold Settings — active threshold configuration
    6. Aggregation Notes — documents the aggregation method used
    """
    output = BytesIO()

    if summary is not None:
        summary_data = [
            {"Campaign Metric": "Total LGAs Analysed", "Value": summary.total_lgas},
            {"Campaign Metric": "Total Wards Analysed", "Value": summary.total_wards},
            {"Campaign Metric": "Total Planned Settlements", "Value": summary.total_settlements},
            {"Campaign Metric": "Visited Settlements", "Value": f"{summary.visited_settlements:,} ({summary.visitation_pct}%)"},
            {"Campaign Metric": "Unvisited Settlements", "Value": f"{summary.total_settlements - summary.visited_settlements:,} ({round(100 - summary.visitation_pct, 1)}%)"},
            {"Campaign Metric": "Avg. Vaccination Coverage %", "Value": f"{summary.avg_vaccination_pct}%"},
            {"Campaign Metric": "Avg. Household Coverage %", "Value": f"{summary.avg_household_pct}%"},
        ]
        if df_issues is not None and not df_issues.empty:
            summary_data.append({"Campaign Metric": "Wards with Operational Issues", "Value": df_issues[COL_WARD].nunique()})
    else:
        summary_data = [{"Campaign Metric": "Status", "Value": "No Summary Available"}]

    df_summary_sheet = pd.DataFrame(summary_data)

    notes_data = [
        {
            "Indicator / Item": "% Visitation (Ward Level)",
            "Aggregation Method": "(count of visited settlements / total settlements in ward) × 100",
        },
        {
            "Indicator / Item": "% Vaccination (Ward Level)",
            "Aggregation Method": "Mean of settlement-level Vaccination Coverage % values",
        },
        {
            "Indicator / Item": "% Household Coverage (Ward Level)",
            "Aggregation Method": "Mean of settlement-level Household Coverage % values",
        },
        {
            "Indicator / Item": "Weighting Note",
            "Aggregation Method": (
                "The mean-of-settlement-percentages approach treats each settlement "
                "equally regardless of population size."
            ),
        },
    ]
    df_notes = pd.DataFrame(notes_data)

    # Build threshold documentation sheet.
    threshold_rows: list[dict] = []
    active_thresholds = thresholds or DEFAULT_THRESHOLDS
    for indicator, threshold in active_thresholds.items():
        threshold_rows.append({
            "Indicator": indicator,
            "🔴 Red": f"Below {threshold.yellow_min}%",
            "🟡 Yellow": f"{threshold.yellow_min}% - {threshold.green_min - 0.1}%",
            "🟢 Green": f"{threshold.green_min}% and above",
        })
    df_thresholds = pd.DataFrame(threshold_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary_sheet.to_excel(writer, sheet_name="Executive Summary", index=False)
        df_settlement.to_excel(writer, sheet_name="Settlement Coverage", index=False)
        df_ward.to_excel(writer, sheet_name="Ward Coverage", index=False)
        if df_issues is not None and not df_issues.empty:
            df_issues.to_excel(writer, sheet_name="Operational Issues", index=False)
        df_thresholds.to_excel(writer, sheet_name="Threshold Settings", index=False)
        df_notes.to_excel(writer, sheet_name="Aggregation Notes", index=False)

        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _apply_excel_formatting(
                ws,
                sheet_name,
                thresholds=active_thresholds,
            )

    return output.getvalue()
